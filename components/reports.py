from fpdf import FPDF
from openpyxl import Workbook
from io import BytesIO
import pandas as pd
import streamlit as st
from components.charts import plot_progress_chart, plot_category_averages


def generate_pdf_report(participant_data: dict, progress_data: pd.DataFrame, stats: dict, averages: dict) -> BytesIO:
    """
    Generiert einen PDF-Bericht für einen Teilnehmer.

    Args:
        participant_data (dict): Daten des Teilnehmers.
        progress_data (pd.DataFrame): Fortschrittsdaten des Teilnehmers.
        stats (dict): Statistiken des Teilnehmers.
        averages (dict): Durchschnittswerte der Kategorien.

    Returns:
        BytesIO: PDF-Bericht im Speicher.
    """
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Titel
    pdf.set_font("Arial", size=16)
    pdf.cell(200, 10, txt="Bericht für Mathematikkurs", ln=True, align="C")

    # Teilnehmerinformationen
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Name: {participant_data['name']}", ln=True)
    pdf.cell(200, 10, txt=f"SV-Nummer: {participant_data['sv_number']}", ln=True)
    pdf.cell(200, 10, txt=f"Eintrittsdatum: {participant_data['entry_date']}", ln=True)
    pdf.cell(200, 10, txt=f"Austrittsdatum: {participant_data['exit_date']}", ln=True)

    # Statistiken
    pdf.ln(10)
    pdf.set_font("Arial", size=14)
    pdf.cell(200, 10, txt="Statistiken", ln=True)
    pdf.set_font("Arial", size=12)
    pdf.cell(200, 10, txt=f"Durchschnitt der letzten zwei Tests: {stats['Durchschnitt_letzte_zwei']}%", ln=True)

    # Durchschnittswerte der Kategorien
    pdf.ln(10)
    pdf.cell(200, 10, txt="Durchschnittswerte der Kategorien:", ln=True)
    for category, avg in averages.items():
        pdf.cell(200, 10, txt=f"{category}: {avg}%", ln=True)

    # Fortschrittsdiagramm einfügen
    pdf.ln(10)
    pdf.cell(200, 10, txt="Fortschrittsdiagramm", ln=True)
    plt_file = BytesIO()
    plot_progress_chart(progress_data)
    plt.savefig(plt_file, format="png")
    plt_file.seek(0)
    pdf.image(plt_file, x=10, y=None, w=190)

    # Rückgabe des PDFs
    pdf_file = BytesIO()
    pdf.output(pdf_file)
    pdf_file.seek(0)
    return pdf_file


def generate_excel_report(participant_data: dict, progress_data: pd.DataFrame, stats: dict, averages: dict) -> BytesIO:
    """
    Generiert einen Excel-Bericht für einen Teilnehmer.

    Args:
        participant_data (dict): Daten des Teilnehmers.
        progress_data (pd.DataFrame): Fortschrittsdaten des Teilnehmers.
        stats (dict): Statistiken des Teilnehmers.
        averages (dict): Durchschnittswerte der Kategorien.

    Returns:
        BytesIO: Excel-Bericht im Speicher.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bericht"

    # Teilnehmerinformationen
    ws.append(["Teilnehmerbericht"])
    ws.append(["Name", participant_data["name"]])
    ws.append(["SV-Nummer", participant_data["sv_number"]])
    ws.append(["Eintrittsdatum", participant_data["entry_date"]])
    ws.append(["Austrittsdatum", participant_data["exit_date"]])
    ws.append([])

    # Statistiken
    ws.append(["Statistiken"])
    ws.append(["Durchschnitt der letzten zwei Tests", stats["Durchschnitt_letzte_zwei"]])
    ws.append([])

    # Durchschnittswerte der Kategorien
    ws.append(["Durchschnittswerte der Kategorien"])
    for category, avg in averages.items():
        ws.append([category, avg])
    ws.append([])

    # Fortschrittsdaten
    ws.append(["Fortschrittsdaten"])
    ws.append(list(progress_data.columns))
    for row in progress_data.itertuples(index=False):
        ws.append(list(row))

    # Rückgabe der Excel-Datei
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file


def download_reports(participant_data: dict, progress_data: pd.DataFrame, stats: dict, averages: dict) -> None:
    """
    Stellt die Berichte als Download zur Verfügung.

    Args:
        participant_data (dict): Teilnehmerdaten.
        progress_data (pd.DataFrame): Fortschrittsdaten.
        stats (dict): Statistiken.
        averages (dict): Durchschnittswerte der Kategorien.

    Returns:
        None
    """
    # PDF-Bericht
    pdf_report = generate_pdf_report(participant_data, progress_data, stats, averages)
    st.download_button(
        label="PDF-Bericht herunterladen",
        data=pdf_report,
        file_name=f"{participant_data['name']}_Bericht.pdf",
        mime="application/pdf",
    )

    # Excel-Bericht
    excel_report = generate_excel_report(participant_data, progress_data, stats, averages)
    st.download_button(
        label="Excel-Bericht herunterladen",
        data=excel_report,
        file_name=f"{participant_data['name']}_Bericht.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
